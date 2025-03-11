import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import rows_from_range
import datetime
import csv
import os

workbook = Workbook()

project_path = "C:/Users/Martina/Desktop/škola/informatika/git.projectinf/"
path = "C:/Users/Martina/Desktop/škola/informatika/git.projectinf/Dummymappe2csv.xlsx"
# Otvírání excel souboru ve formátu "robot"
print(f"Please select a file in format: {project_path}Name.xlsx")
# path = filename = input()
wb_obj = openpyxl.load_workbook(path, data_only=True)
sheet_obj = wb_obj.active

row_number = sheet_obj.max_row
column_number = sheet_obj.max_column

l_fall30_vers = ["AN Arbeitslosenve",
                 "AN Krankenvers.",
                 "AN Plegeversich.",
                 "AN Renteversich."]

l_fall_27 = ["AG Pauschsteuer",
             "AN Pauschsteuer",
             "Kirchensteuer",
             "Lohnsteuer"]

# Textový soubor s Lohnarty jistými pro fall30
with open(f"{project_path}Fall30.txt", "r") as fall_30:
    lines_from_text = fall_30.readlines()
    for i in range(0, len(lines_from_text)):
        lines_fall_30 = lines_from_text[i].replace("\n", "")
        lines_fall_30 = [line.strip() for line in lines_from_text if line.strip()]

# Převedení excelu na csv pro snazší použití později
with open(f"{project_path}Dummymappe2csv.csv", "w", newline="") as file_handle:
    csv_writer = csv.writer(file_handle, delimiter=";")
    for row in sheet_obj.iter_rows(min_row=2):
        csv_writer.writerow([cell.value for cell in row])

# Vytvoření dictionary ze všech lidí v dokumentu
with open(f"{project_path}Dummymappe2csv.csv", "r", encoding="latin-1", newline="") as f:
    csv_rows = list(csv.reader(f, delimiter=';'))
    people_dict = {}
    for line in csv_rows:
        if line[1] not in people_dict:
            people_dict[line[1]] = [line]
        else:
            people_dict[line[1]].append(line)

all_the_people = list(people_dict.keys())


# Třída každého člověka s důležitým info
class Person:
    def __init__(self, i_in_people):
        current_person = all_the_people[i_in_people]
        self.abrk = people_dict[current_person][0][0]
        self.name = people_dict[current_person][0][2]
        self.PN = people_dict[current_person][0][1]
        self.month = []
        self.lohn = []
        self.fall30 = False
        self.fall27 = False


def people_classes(x):
    # Počítání a zapisování měsíců, ve krerých byla změna
    person = Person(x)
    local_month = []
    local_lohn = []
    for i in people_dict[all_the_people[x]]:
        local_month.append(i[3])
        local_lohn.append(i[11])
    local_month = list(dict.fromkeys(local_month))
    local_lohn = [i[11].strip() for i in people_dict[all_the_people[x]] if i[11].strip()]

    # Loop skrz list s jistým fall 30
    person.fall30 = any(k in lines_fall_30 for k in local_lohn)
    print(local_lohn, person.fall30)

    # Loop skrz list s versicherung a procenta
    if person.fall30 is False:
        if any(k in l_fall30_vers for k in local_lohn) and not i[13] == "":
            person.fall30 = True
            print(local_lohn, person.fall30)

    # Loop skrz list s steuerung a procenta
    if person.fall30 is False:
        if any(k in l_fall_27 for k in local_lohn) and not i[13] == "":
            person.fall27 = True
            print(local_lohn, person.fall30, person.fall27)

    person.month = local_month
    person.lohn = local_lohn
    return person


# Loop skrz všechny lidi
list_of_People = []
for i in range(len(all_the_people)):
    list_of_People.append(people_classes(i))


def not_valid():
    print("Invalid input given")
    if i == 4:
        print("Invalid input given five times - Program ends.")
        exit()


# Název výsledného souboru
for i in range(5):
    print("Are you making Qualität for this month write: Y/n")
    answer = input()
    # Pojmenování excelu podle aktuálního měsíce
    if answer == 'Y':
        month_tuple = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
        last_month = month_tuple.strftime("%m")
        current_year = month_tuple.strftime("%y")
        break
    elif answer == 'n':
        print("Input the month of the given data in format '01'.")
        last_month = input()
        try:
            month_input = int(last_month)
        except ValueError:
            not_valid()
        else:
            print("Input the year of the given data in format '25'.")
            current_year = input()
            try:
                year_input = int(current_year)
            except ValueError:
                not_valid()
    else:
        not_valid()

end_name = (f"{project_path}Qualität_{last_month}_{current_year}.xlsx")


# Vytvoření výsledného excelu
def final_report():
    sheet = workbook.active
    asrow = 1000+len(all_the_people)

    # Nutná legenda na konci výsledného excelu
    file_name_source = f'{project_path}End_of_Report.xlsx'
    wb_source = openpyxl.load_workbook(file_name_source)
    sheet_source = wb_source.active

    range_str = 'A1:R64'

    for row in rows_from_range(range_str):
        for cell in row:
            dest_sheet_cell = sheet[cell].offset(row=asrow-1)
            source_cell = sheet_source[cell]

            merged = False
            for merged_range in sheet_source.merged_cells.ranges:
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                max_col = merged_range.max_col
                max_row = merged_range.max_row
                if min_row <= source_cell.row <= max_row and min_col <= source_cell.column <= max_col:
                    merged = True
                    if (source_cell.row == min_row) and (source_cell.column == min_col):
                        m_source = sheet_source.cell(row=min_row, column=min_col)
                        m_dest = sheet[m_source.coordinate].offset(row=asrow-1)

                        m_dest.value = m_source.value
                        m_dest.font = Font(bold=m_source.font.bold,
                                           color=m_source.font.color)

                        sheet.merge_cells(start_row=m_dest.row,
                                          start_column=m_dest.column,
                                          end_row=m_dest.row + max_row - min_row,
                                          end_column=m_dest.column + max_col - min_col)
                    break

            if not merged:
                dest_sheet_cell.value = source_cell.value
                dest_sheet_cell.font = Font(bold=source_cell.font.bold,
                                            color=source_cell.font.color)
    # Gesamtergebnis hodnoty
    for i in range(0, 12):
        row_gsmterg = str(asrow)
        column_gsmterg = chr(64 + 2 + i)
        sheet[column_gsmterg + row_gsmterg] = f'=SUM({column_gsmterg}2:{column_gsmterg}{str(len(all_the_people)+1)})'

    cl = sheet.cell(row=asrow+11, column=3)
    cl.value = f'=COUNT(B2:M{str(asrow-1)})'

    # Anzahl hodnoty
    for i in range(0, 9):
        row_lngth = len(all_the_people)
        row_strmln = str(asrow + 15 + i)
        sheet['C' + row_strmln] = f'=SUMIF(N2:N{str(row_lngth)}, B{row_strmln}, O2:O{str(row_lngth+1)})'

    for i in range(0, 9):
        row_strmln = str(asrow + 27 + i)
        sheet['C' + row_strmln] = f'=SUMIF(N2:N{str(row_lngth)}, B{row_strmln}, O2:O{str(row_lngth+1)})'

    # Nadepsání tabulky s lidmi
    c = sheet['A1']
    c.value = "Zeilenbeschriftungen"
    sheet.column_dimensions['A'].width = 20
    c1 = sheet['N1']
    c1.value = "RR A."
    c2 = sheet['O1']
    c2.value = "Grund"

    # Formátování 12 měsíců do aktuálního měsíce
    for i in range(0, 12):
        sheet_cell = sheet.cell(row=1, column=13 - i)
        l_m_int = int(last_month)
        c_y_int = int(current_year)
        if l_m_int - i >= 1:
            sheet_cell.value = str(l_m_int - i) + "/" + current_year
        else:
            sheet_cell.value = str((l_m_int + 12) - i) + "/" + str(c_y_int - 1)

    # Formátování získaných dat/výsledků
    for i in range(len(all_the_people)):
        person = list_of_People[i]
        sheet_cell = sheet.cell(row=i+2, column=1)
        sheet_cell.value = (person.abrk + "/"
                            + person.PN + "/" + person.name)

        if person.fall30:
            sheet_cell1 = sheet.cell(row=i+2, column=15)
            sheet_cell1.value = 30

        if person.fall27:
            sheet_cell1 = sheet.cell(row=i+2, column=15)
            sheet_cell1.value = 27

        for k in range(0, len(person.month)):
            month_position = int(last_month) - int(person.month[k])
            if month_position >= 0:
                sheet_cell = sheet.cell(row=i+2, column=13-month_position)
                sheet_cell.value = 1
            else:
                sheet_cell = sheet.cell(row=i+2, column=1-month_position)
                sheet_cell.value = 1
        row_sum = str(i + 2)
        sheet["N" + row_sum] = f'=SUM(B{row_sum}:M{row_sum})'

    workbook.save(filename=end_name)


final_report()
os.remove(f"{project_path}Dummymappe2csv.csv")
